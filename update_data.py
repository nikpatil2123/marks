#!/usr/bin/env python3
"""
Run this script whenever you update 'certificates data.xlsx' to sync ALL_DATA in index.html.
Usage: python3 update_data.py
"""
import json, re, sys
import openpyxl
from datetime import datetime

XLSX = 'certificates data.xlsx'
HTML = 'index.html'
MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def fmt_date(val):
    if isinstance(val, datetime):
        return f"{val.day:02d} {MONTHS[val.month-1]} {val.year}"
    return str(val).strip() if val else ""

def is_ok(status):
    return bool(status and 'Document successfully created' in str(status))

def s(v):
    return str(v).strip() if v else ""

try:
    wb = openpyxl.load_workbook(XLSX)
except FileNotFoundError:
    print(f"ERROR: '{XLSX}' not found. Run this script from the marks/ folder.")
    sys.exit(1)

all_data = []

# ── Sheet 1: 1 name ─────────────────────────────────────────────
ws = wb['1']
for row in ws.iter_rows(min_row=2, values_only=True):
    row = (list(row) + [None]*9)[:9]
    program, name, subject, srno, date, doc_id, url, hyperlink, status = row
    if not srno and not name:
        continue
    all_data.append({
        "sheetName": "1",
        "sheetLabel": "Sheet 1",
        "marksEach": 9,
        "programName": s(program),
        "srno": s(srno),
        "names": [s(name)] if name else [],
        "subject": s(subject),
        "date": fmt_date(date),
        "docName": f"{s(name)}-{s(subject)}" if name and subject else s(name or subject),
        "link": s(url),
        "ok": is_ok(status),
    })

# ── Sheet 2: 2 names ─────────────────────────────────────────────
ws = wb['2']
for row in ws.iter_rows(min_row=2, values_only=True):
    row = (list(row) + [None]*12)[:12]
    program, name1, name2, subject, srno, date, doc_id, url, hyperlink, status = row[:10]
    if not srno and not name1:
        continue
    names = [s(n) for n in [name1, name2] if n and s(n)]
    all_data.append({
        "sheetName": "2",
        "sheetLabel": "Sheet 2",
        "marksEach": 4,
        "programName": s(program),
        "srno": s(srno),
        "names": names,
        "subject": s(subject),
        "date": fmt_date(date),
        "docName": s(subject),
        "link": s(url),
        "ok": is_ok(status),
    })

# ── Sheet 3: 3 names ─────────────────────────────────────────────
ws = wb['3']
for row in ws.iter_rows(min_row=2, values_only=True):
    row = (list(row) + [None]*11)[:11]
    program, name1, name2, name3, subject, srno, date, doc_id, url, hyperlink, status = row[:11]
    if not srno and not name1:
        continue
    names = [s(n) for n in [name1, name2, name3] if n and s(n)]
    doc_name = "-".join(names) + (f"-{s(subject)}" if subject else "")
    all_data.append({
        "sheetName": "3",
        "sheetLabel": "Sheet 3",
        "marksEach": 3,
        "programName": s(program),
        "srno": s(srno),
        "names": names,
        "subject": s(subject),
        "date": fmt_date(date),
        "docName": doc_name,
        "link": s(url),
        "ok": is_ok(status),
    })

print(f"Loaded: Sheet 1={sum(1 for d in all_data if d['sheetName']=='1')}, "
      f"Sheet 2={sum(1 for d in all_data if d['sheetName']=='2')}, "
      f"Sheet 3={sum(1 for d in all_data if d['sheetName']=='3')}, "
      f"Total={len(all_data)}")

# ── Patch index.html ─────────────────────────────────────────────
with open(HTML, 'r', encoding='utf-8') as f:
    content = f.read()

new_json = json.dumps(all_data, ensure_ascii=False, separators=(',', ':'))
new_content, n = re.subn(
    r'const ALL_DATA = \[.*?\];',
    f'const ALL_DATA = {new_json};',
    content,
    flags=re.DOTALL
)

if n == 0:
    print("ERROR: Could not find ALL_DATA in index.html")
    sys.exit(1)

with open(HTML, 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f"Done! index.html updated with {len(all_data)} records.")
print("Refresh the browser to see changes.")
